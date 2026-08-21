"""Convert between the Univer snapshot and ``.xlsx`` / ``.csv``.

The snapshot (``IWorkbookData``) is LazyClaw's source of truth; ``.xlsx`` is a
derived format produced on export and parsed on import. We write formulas as
text (``=SUM(A1:A2)``) so Excel/Google Sheets recompute them on open — see
:mod:`lazyclaw.sheets.recalc` for the server-side eval used when the agent
edits formulas without a browser pass.

openpyxl (MIT) handles both directions and preserves formulas; we read with
``data_only=False`` so formula text survives the round-trip.
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime, time
from typing import Any
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from lazyclaw.sheets import snapshot as S
from lazyclaw.sheets import xlsx_styles as XS

# Excel caps worksheet names at 31 chars.
_SHEET_NAME_MAX = 31


def _native_for_xlsx(value: Any) -> Any:
    """Coerce a cell value to something openpyxl can write directly."""
    if value is None or isinstance(value, (int, float, str, bool, datetime, date, time)):
        return value
    return str(value)


def snapshot_to_xlsx(snap: dict[str, Any]) -> bytes:
    """Render a Univer snapshot to ``.xlsx`` bytes (formulas kept as text)."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet; we recreate in order

    sheets = snap.get("sheets") or {}
    order = snap.get("sheetOrder") or list(sheets.keys())
    if not order:
        wb.create_sheet("Sheet1")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    registry = snap.get("styles") or {}

    for sid in order:
        sdata = sheets.get(sid, {})
        ws = wb.create_sheet(title=(sdata.get("name") or "Sheet")[:_SHEET_NAME_MAX])
        cell_data = sdata.get("cellData") or {}
        for r_key, cols in cell_data.items():
            for c_key, cell in (cols or {}).items():
                if not cell:
                    continue
                r, c = int(r_key) + 1, int(c_key) + 1
                formula = cell.get("f")
                if formula:
                    target = ws.cell(
                        row=r, column=c,
                        value=formula if formula.startswith("=") else f"={formula}",
                    )
                elif cell.get("v") is not None:
                    target = ws.cell(row=r, column=c, value=_native_for_xlsx(cell["v"]))
                else:
                    # Style-only cell — still needs to exist to carry formatting.
                    target = ws.cell(row=r, column=c)
                XS.apply_to_cell(target, _style_of(registry, cell.get("s")))

        _write_geometry(ws, sdata)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _style_of(registry: dict[str, Any], ref: Any) -> dict[str, Any]:
    """``ICellData.s`` → its style dict (it may be an inline dict or an id)."""
    if isinstance(ref, dict):
        return ref
    if isinstance(ref, str):
        entry = registry.get(ref)
        return entry if isinstance(entry, dict) else {}
    return {}


def _write_geometry(ws: Any, sdata: dict[str, Any]) -> None:
    """Column widths, row heights, merges and the freeze pane.

    Merges go LAST: openpyxl replaces every covered cell with a ``MergedCell``,
    which raises on assignment — so this must run after all the cell writes.
    """
    for c_key, entry in (sdata.get("columnData") or {}).items():
        width = (entry or {}).get("w")
        if isinstance(width, (int, float)):
            letter = S.col_to_letter(int(c_key))
            ws.column_dimensions[letter].width = XS.px_to_char_width(width)

    for r_key, entry in (sdata.get("rowData") or {}).items():
        height = (entry or {}).get("h")
        if isinstance(height, (int, float)):
            ws.row_dimensions[int(r_key) + 1].height = XS.px_to_points(height)

    freeze = sdata.get("freeze")
    if isinstance(freeze, dict):
        rows = int(freeze.get("ySplit") or 0)
        cols = int(freeze.get("xSplit") or 0)
        if rows or cols:
            ws.freeze_panes = S.rc_to_a1(rows, cols)

    for rect in sdata.get("mergeData") or []:
        if not isinstance(rect, dict):
            continue
        try:
            ws.merge_cells(
                start_row=int(rect["startRow"]) + 1,
                start_column=int(rect["startColumn"]) + 1,
                end_row=int(rect["endRow"]) + 1,
                end_column=int(rect["endColumn"]) + 1,
            )
        except (KeyError, TypeError, ValueError):
            continue


def xlsx_to_snapshot(data: bytes, name: str | None = None) -> dict[str, Any]:
    """Parse ``.xlsx`` bytes into a Univer snapshot (formulas preserved)."""
    wb = load_workbook(io.BytesIO(data), data_only=False)
    snap = S.blank_workbook(name or "Imported")
    snap["sheets"] = {}
    snap["sheetOrder"] = []

    for ws in wb.worksheets:
        sid = f"sh-{uuid4().hex[:12]}"
        cell_data: dict[str, dict[str, Any]] = {}
        max_r = max_c = 0
        for row in ws.iter_rows():
            for cell in row:
                style_id = XS.intern(snap, XS.from_cell(cell))
                if cell.value is None:
                    # A cell with formatting but no value is still worth
                    # keeping — it's a styled empty header or a banner.
                    if style_id:
                        r, c = cell.row - 1, cell.column - 1
                        cell_data.setdefault(str(r), {})[str(c)] = {"s": style_id}
                    continue
                r, c = cell.row - 1, cell.column - 1
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    cd: dict[str, Any] = {"f": v}
                elif isinstance(v, (datetime, date, time)):
                    cd = {"v": v.isoformat()}
                else:
                    cd = {"v": v}
                if style_id:
                    cd["s"] = style_id
                cell_data.setdefault(str(r), {})[str(c)] = cd
                max_r, max_c = max(max_r, r), max(max_c, c)
        sheet_obj: dict[str, Any] = {
            "id": sid,
            "name": ws.title,
            "rowCount": max(max_r + 1, S.DEFAULT_ROWS),
            "columnCount": max(max_c + 1, S.DEFAULT_COLS),
            "cellData": cell_data,
        }
        _read_geometry(ws, sheet_obj, max_c)
        snap["sheets"][sid] = sheet_obj
        snap["sheetOrder"].append(sid)

    if not snap["sheetOrder"]:
        return S.blank_workbook(name or "Imported")
    return snap


def _read_geometry(ws: Any, sheet_obj: dict[str, Any], max_col: int) -> None:
    """Column widths, row heights, merges and the freeze pane, from a workbook.

    Only ``customWidth`` dimensions are honoured, and a span is expanded no
    further than the used columns: an Excel-authored file routinely declares
    one dimension covering ``min=1, max=16384``, which taken literally would
    write sixteen thousand keys into the encrypted blob. Note also that
    *indexing* ``ws.column_dimensions['A']`` CREATES an entry, so this iterates
    ``.items()`` and never indexes.
    """
    column_data: dict[str, dict[str, Any]] = {}
    for _, dim in ws.column_dimensions.items():
        width = getattr(dim, "width", None)
        if width is None or not getattr(dim, "customWidth", False):
            continue
        lo = int(getattr(dim, "min", 1) or 1)
        hi = int(getattr(dim, "max", lo) or lo)
        hi = min(hi, lo + XS.MAX_COL_SPAN - 1, max(max_col + 1, lo))
        for index in range(lo - 1, hi):
            column_data[str(index)] = {"w": XS.char_width_to_px(width)}
    if column_data:
        sheet_obj["columnData"] = column_data

    row_data: dict[str, dict[str, Any]] = {}
    for index, dim in ws.row_dimensions.items():
        height = getattr(dim, "height", None)
        if height is None or int(index) < 1:
            continue
        row_data[str(int(index) - 1)] = {"h": XS.points_to_px(height)}
    if row_data:
        sheet_obj["rowData"] = row_data

    merged = getattr(ws, "merged_cells", None)
    merges = [
        {
            "startRow": rng.min_row - 1, "startColumn": rng.min_col - 1,
            "endRow": rng.max_row - 1, "endColumn": rng.max_col - 1,
        }
        for rng in (getattr(merged, "ranges", None) or [])
    ]
    if merges:
        sheet_obj["mergeData"] = merges

    frozen = getattr(ws, "freeze_panes", None)
    if isinstance(frozen, str) and frozen:
        try:
            row, col = S.a1_to_rc(frozen)
        except ValueError:
            return
        if row or col:
            sheet_obj["freeze"] = {
                "xSplit": col, "ySplit": row, "startRow": row, "startColumn": col,
            }


def snapshot_to_csv(snap: dict[str, Any], sheet: int | str = 0) -> str:
    """Render one worksheet's display grid to CSV text."""
    grid = S.as_grid(snap, sheet)
    out = io.StringIO()
    writer = csv.writer(out)
    for row in grid:
        writer.writerow(row)
    return out.getvalue()
