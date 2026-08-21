"""Tests for snapshot ⇄ xlsx/csv conversion (lazyclaw/sheets/xlsx_io.py)."""

from __future__ import annotations

import io

import pytest

from lazyclaw.sheets import snapshot as S
from lazyclaw.sheets.xlsx_io import (
    snapshot_to_csv,
    snapshot_to_xlsx,
    xlsx_to_snapshot,
)


def _budget():
    snap = S.blank_workbook("Budget", workbook_id="wb", sheet_id="sh", sheet_name="Sheet1")
    return S.set_cells(snap, [
        {"row": 0, "col": 0, "value": 10},
        {"row": 1, "col": 0, "value": 20},
        {"row": 2, "col": 0, "formula": "=SUM(A1:A2)"},
        {"row": 0, "col": 1, "value": "hello"},
    ])


def test_xlsx_roundtrip_preserves_values_and_formulas():
    xb = snapshot_to_xlsx(_budget())
    assert isinstance(xb, bytes) and len(xb) > 0

    back = xlsx_to_snapshot(xb, name="Budget")
    assert S.sheet_names(back) == ["Sheet1"]
    assert S.cell_display(S.get_cell(back, 0, 0)) == 10
    assert S.cell_display(S.get_cell(back, 0, 1)) == "hello"
    # formula text survives the round-trip (data_only=False)
    assert S.get_cell(back, 2, 0)["f"] == "=SUM(A1:A2)"


def test_xlsx_formula_gets_leading_equals_on_export():
    snap = S.set_cell(
        S.blank_workbook("T", workbook_id="wb", sheet_id="sh"),
        0, 0, formula="A1+1",
    )
    # snapshot already normalises to "=A1+1"; ensure it survives to xlsx + back
    back = xlsx_to_snapshot(snapshot_to_xlsx(snap))
    assert S.get_cell(back, 0, 0)["f"].startswith("=")


def test_multi_sheet_roundtrip():
    snap = S.blank_workbook("Multi", workbook_id="wb", sheet_id="sh1", sheet_name="First")
    # add a second worksheet manually
    snap["sheets"]["sh2"] = {
        "id": "sh2", "name": "Second", "rowCount": 1000, "columnCount": 20,
        "cellData": {"0": {"0": {"v": 42}}},
    }
    snap["sheetOrder"].append("sh2")
    snap = S.set_cell(snap, 0, 0, value=1, sheet=0)

    back = xlsx_to_snapshot(snapshot_to_xlsx(snap))
    assert S.sheet_names(back) == ["First", "Second"]
    assert S.cell_display(S.get_cell(back, 0, 0, sheet="Second")) == 42


def test_empty_snapshot_exports_and_reimports():
    blank = S.blank_workbook("Empty", workbook_id="wb", sheet_id="sh")
    back = xlsx_to_snapshot(snapshot_to_xlsx(blank))
    assert back["sheetOrder"]  # at least one sheet


def test_csv_export_uses_display_grid():
    csv_text = snapshot_to_csv(_budget())
    lines = csv_text.strip().splitlines()
    # grid is dense to 2 cols because B1 has "hello"
    assert lines[0] == "10,hello"
    assert lines[1] == "20,"
    # no cached value on the formula cell → shows the formula text (+ empty B)
    assert lines[2] == "=SUM(A1:A2),"


# ───────────── styles + geometry round-trip (2026-08-21) ────────────


def _styled():
    from lazyclaw.sheets import geometry as G
    from lazyclaw.sheets import styles as ST

    wb = S.blank_workbook("T", workbook_id="wb", sheet_id="sh")
    wb = S.set_cells(wb, [
        {"cell": "A1", "value": "Item"}, {"cell": "B1", "value": "Cost"},
        {"cell": "A2", "value": "Rent"}, {"cell": "B2", "value": 1200},
        {"cell": "B3", "formula": "=SUM(B2:B2)"},
    ])
    wb = ST.apply_style_a1(wb, "A1:B1", {
        "bold": True, "bg": "#E8E8E8", "align": "center",
    })
    wb = ST.apply_style_a1(wb, "B2:B3", {"number_format": "currency"})
    wb = ST.apply_style_a1(wb, "A2", {"italic": True, "color": "#FF0000"})
    wb = G.set_column_width(wb, 0, 160)
    wb = G.set_row_height(wb, 0, 32)
    wb = G.merge_cells(wb, 5, 0, 5, 2)
    return G.freeze_panes(wb, rows=1)


def test_xlsx_roundtrip_preserves_styles():
    from lazyclaw.sheets import styles as ST

    back = xlsx_to_snapshot(snapshot_to_xlsx(_styled()), "RT")
    assert ST.get_style_view(back, 0, 0) == {
        "bold": True, "bg": "#E8E8E8", "align": "center",
    }
    assert ST.get_style_view(back, 1, 0) == {"italic": True, "color": "#FF0000"}
    assert ST.resolve_style(back, 1, 1)["n"]["pattern"] == "$#,##0.00"


def test_xlsx_roundtrip_preserves_geometry():
    from lazyclaw.sheets import geometry as G

    back = xlsx_to_snapshot(snapshot_to_xlsx(_styled()), "RT")
    assert G.get_column_width(back, 0) == pytest.approx(160, abs=1)
    assert G.get_row_height(back, 0) == pytest.approx(32, abs=1)
    assert G.merged_range_at(back, 5, 1) == {
        "startRow": 5, "startColumn": 0, "endRow": 5, "endColumn": 2
    }
    assert G.frozen_counts(back) == (1, 0)


def test_xlsx_roundtrip_still_preserves_values_and_formulas():
    back = xlsx_to_snapshot(snapshot_to_xlsx(_styled()), "RT")
    assert S.cell_display(S.get_cell(back, 1, 1)) == 1200
    assert S.get_cell(back, 2, 1)["f"] == "=SUM(B2:B2)"


def test_xlsx_import_skips_theme_colours_and_the_general_format():
    """A plain cell's font colour comes back as theme/rgb=None, and its number
    format as 'General'. Without both guards every imported cell would acquire
    a garbage explicit colour and a bogus n:{pattern:'General'}."""
    plain = S.set_cells(
        S.blank_workbook("P", workbook_id="wb", sheet_id="sh"),
        [{"cell": "A1", "value": "hello"}],
    )
    back = xlsx_to_snapshot(snapshot_to_xlsx(plain), "P")
    assert back["styles"] == {}
    assert "s" not in S.get_cell(back, 0, 0)


def test_xlsx_import_does_not_explode_a_full_width_column_span():
    """Excel files declare one dimension covering min=1..max=16384."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws.column_dimensions["A"].min = 1
    ws.column_dimensions["A"].max = 16384
    # `customWidth` is derived from `width` being set — it has no setter.
    ws.column_dimensions["A"].width = 12
    buf = io.BytesIO()
    wb.save(buf)

    back = xlsx_to_snapshot(buf.getvalue(), "Wide")
    column_data = back["sheets"][back["sheetOrder"][0]].get("columnData") or {}
    assert len(column_data) <= 256, "a full-sheet span was expanded verbatim"


def test_style_only_cells_survive_the_roundtrip():
    from lazyclaw.sheets import styles as ST

    wb = ST.apply_style_a1(
        S.blank_workbook("S", workbook_id="wb", sheet_id="sh"),
        "A1", {"bg": "yellow"},
    )
    back = xlsx_to_snapshot(snapshot_to_xlsx(wb), "S")
    assert ST.get_style_view(back, 0, 0) == {"bg": "#FFFF00"}
